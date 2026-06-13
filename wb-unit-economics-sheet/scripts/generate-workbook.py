#!/usr/bin/env python3
"""Генерирует таблицу юнит-экономики WB для Google Таблиц и Excel."""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
LAST_ROW = 500

HEADERS = [
    "№", "Артикул ВБ", "Артикул продавца", "Бренд", "Название",
    "Остаток FBO", "Остаток FBS", "Остаток поставщика", "Комментарий", "Заказы 7д",
    "Цена закупки", "Цена продажи", "Цена базовая", "% скидки",
    "Цена продажи кальк.", "Прибыль FBO кальк.", "Рентаб. FBO кальк.%",
    "Прибыль FBS кальк.", "Рентаб. FBS кальк.%",
    "Наша цена на ВБ", "СПП",
    "% Выкупа", "Длина", "Ширина", "Высота", "Объём", "Коэфф. склада",
    "Доставка базовая", "Доставка с возвратом",
    "Налог УСН %", "Налог ₽", "Доп. комиссия %",
    "Комиссия FBO кат.%", "Комиссия FBO итог%", "Комиссия FBO ₽",
    "Комиссия FBS кат.%", "Комиссия FBS итог%", "Комиссия FBS ₽",
    "Упаковка", "Хранение", "Брак %", "Брак ₽",
    "Прибыль FBO", "Маржа FBO %", "Рентаб. FBO %",
    "Прибыль FBS", "Маржа FBS %", "Рентаб. FBS %",
]


def formulas_for_row(r: int, google: bool = False) -> dict[int, str]:
    """Формулы строки r. google=True → разделитель ; и десятичная , (локаль RU)."""
    pack = f"IF(AM{r}>0;AM{r};'_Настройки'!$B$9)" if google else f"IF(AM{r}>0,AM{r},'_Настройки'!$B$9)"
    coeff = f"IF(AA{r}>0;AA{r};'_Настройки'!$B$10)" if google else f"IF(AA{r}>0,AA{r},'_Настройки'!$B$10)"
    sep = ";" if google else ","
    empty = '""'
    dec = lambda v: str(v).replace(".", ",") if google else str(v)

    purchase = (
        f"IFERROR(INDEX('_Цены_закупки'!$B:$B;MATCH(C{r};'_Цены_закупки'!$A:$A;0));{empty})"
        if google
        else f"=IFERROR(INDEX('_Цены_закупки'!$B:$B,MATCH(C{r},'_Цены_закупки'!$A:$A,0)),{empty})"
    )
    fbo_cat = (
        f"IFERROR(INDEX('_Комиссия_ВБ'!$B:$B;MATCH(C{r};'_Комиссия_ВБ'!$A:$A;0));{dec(0.245)})"
        if google
        else f"IFERROR(INDEX('_Комиссия_ВБ'!$B:$B,MATCH(C{r},'_Комиссия_ВБ'!$A:$A,0)),{dec(0.245)})"
    )
    fbs_cat = (
        f"IFERROR(INDEX('_Комиссия_ВБ'!$C:$C;MATCH(C{r};'_Комиссия_ВБ'!$A:$A;0));{dec(0.28)})"
        if google
        else f"IFERROR(INDEX('_Комиссия_ВБ'!$C:$C,MATCH(C{r},'_Комиссия_ВБ'!$A:$A,0)),{dec(0.28)})"
    )

    def F(expr: str) -> str:
        return expr if expr.startswith("=") else f"={expr}"

    sub_rate = f"IF(Z{r}<=0.2,23,IF(Z{r}<=0.4,26,IF(Z{r}<=0.6,29,IF(Z{r}<=0.8,30,32))))"

    raw = {
        1: f"ROW()-1",
        11: purchase.lstrip("="),
        14: f"IF(AND(M{r}>0{sep}L{r}>0){sep}1-L{r}/M{r}{sep}{empty})",
        15: f"IF(OR(Q{r}={empty}{sep}K{r}=0){sep}{empty}{sep}(K{r}+{pack}+AP{r}+AC{r}+Q{r}*K{r})/(1-AH{r}-AD{r}))",
        16: f"IF(O{r}={empty}{sep}{empty}{sep}O{r}-K{r}-O{r}*AH{r}-O{r}*AD{r}-{pack}-AP{r}-AC{r})",
        17: f"IF(AND(O{r}>0{sep}K{r}>0){sep}P{r}/K{r}{sep}{empty})",
        18: f"IF(O{r}={empty}{sep}{empty}{sep}O{r}-K{r}-O{r}*AK{r}-O{r}*AD{r}-{pack}-AP{r}-AC{r})",
        19: f"IF(AND(O{r}>0{sep}K{r}>0){sep}R{r}/K{r}{sep}{empty})",
        21: f"IF(AND(L{r}>0{sep}T{r}>0){sep}1-T{r}/L{r}{sep}{empty})",
        22: f"'_Настройки'!$B$4",
        26: f"IF(AND(W{r}>0{sep}X{r}>0{sep}Y{r}>0){sep}W{r}*X{r}*Y{r}/1000{sep}{empty})",
        28: f"IF(Z{r}>1,('_Настройки'!$B$6+MAX(0{sep}Z{r}-1)*'_Настройки'!$B$7)*{coeff},({sub_rate})*{coeff})",
        29: f"AB{r}*(1+'_Настройки'!$B$8)",
        30: f"'_Настройки'!$B$2",
        31: f"L{r}*AD{r}",
        32: f"'_Настройки'!$B$3",
        33: fbo_cat.lstrip("="),
        34: f"AG{r}+AF{r}",
        35: f"L{r}*AH{r}",
        36: fbs_cat.lstrip("="),
        37: f"AJ{r}+AF{r}",
        38: f"L{r}*AK{r}",
        41: f"'_Настройки'!$B$5",
        42: f"IF(K{r}>0{sep}K{r}*AO{r}{sep}0)",
        43: f"L{r}-K{r}-AI{r}-AE{r}-{pack}-AP{r}-AC{r}",
        44: f"IF(L{r}>0{sep}AQ{r}/L{r}{sep}{empty})",
        45: f"IF(K{r}>0{sep}AQ{r}/K{r}{sep}{empty})",
        46: f"L{r}-K{r}-AL{r}-AE{r}-{pack}-AP{r}-AC{r}",
        47: f"IF(L{r}>0{sep}AT{r}/L{r}{sep}{empty})",
        48: f"IF(K{r}>0{sep}AT{r}/K{r}{sep}{empty})",
    }
    return {col: F(expr) for col, expr in raw.items()}


def setup_settings(ws, google: bool):
    rows = [
        ["Параметр", "Значение", "Описание"],
        ["Налог УСН Доходы", 0.11, "Доля от цены продажи"],
        ["Доп. комиссия WB", 0.0175, "К категорийной комиссии"],
        ["% выкупа", 0.9, "Справочно"],
        ["Брак/потери", 0.01, "Доля от закупки"],
        ["Логистика: 1-й литр, ₽", 46, "Тариф WB"],
        ["Логистика: доп. литр, ₽", 14, "За литр сверх 1"],
        ["Наценка обратной логистики", 0.0454, "К базовой доставке"],
        ["Упаковка по умолчанию, ₽", 65, "Если «Упаковка» пуста"],
        ["Коэфф. склада по умолчанию", 2.2, "Если «Коэфф.» пуст"],
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["C"].width = 36
    if google:
        ws["B2"].number_format = "0,00%"
        ws["B3"].number_format = "0,00%"
        ws["B4"].number_format = "0,00%"
        ws["B5"].number_format = "0,00%"
        ws["B8"].number_format = "0,00%"


def load_json(name):
    path = DATA / name
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def setup_purchases(ws):
    ws.append(["Артикул продавца", "Цена закупки"])
    for art, price in load_json("seed-purchases.json").items():
        ws.append([str(art), float(price)])


def setup_commissions(ws):
    ws.append(["Артикул продавца", "Комиссия FBO", "Комиссия FBS"])
    for art, c in load_json("seed-commissions.json").items():
        ws.append([str(art), float(c["fboCategory"]), float(c["fbsCategory"])])
    ws["B2"].number_format = "0.00%"
    ws["C2"].number_format = "0.00%"


def setup_main(ws, google: bool):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D6A4F")
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 42

    for row in range(2, LAST_ROW + 1):
        for col, formula in formulas_for_row(row, google=google).items():
            ws.cell(row, col, formula)

    money_cols = [11, 12, 13, 15, 16, 18, 20, 28, 29, 31, 35, 38, 39, 42, 43, 46]
    pct_cols = [14, 17, 19, 21, 22, 30, 32, 33, 34, 36, 37, 41, 44, 45, 47, 48]
    for row in range(2, LAST_ROW + 1):
        for c in money_cols:
            ws.cell(row, c).number_format = "#,##0.00"
        for c in pct_cols:
            ws.cell(row, c).number_format = "0.00%"
        ws.cell(row, 26).number_format = "0.00"

    for col, w in {2: 14, 3: 16, 5: 32, 12: 12, 20: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_workbook(google: bool):
    wb = Workbook()
    wb.remove(wb.active)
    setup_settings(wb.create_sheet("_Настройки"), google)
    setup_purchases(wb.create_sheet("_Цены_закупки"))
    setup_commissions(wb.create_sheet("_Комиссия_ВБ"))
    setup_main(wb.create_sheet("Юнитка", 0), google)
    return wb


def main():
    out_google = ROOT / "Юнитка-WB-Google.xlsx"
    out_excel = ROOT / "Юнитка-WB.xlsx"

    build_workbook(google=True).save(out_google)
    build_workbook(google=False).save(out_excel)

    print(f"Google (RU формулы): {out_google}")
    print(f"Excel:               {out_excel}")
    print(f"Строк: {LAST_ROW - 1}")


if __name__ == "__main__":
    main()
